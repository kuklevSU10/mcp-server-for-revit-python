# -*- coding: utf-8 -*-
"""VOR vs BIM comparison — semantic matching via global_patterns.json."""
import json
import os
from mcp.server.fastmcp import Context
from ._constants import CATEGORY_REGISTRY, FT3_TO_M3, FT2_TO_M2
from ._validation import validate_vor_data
from ._patterns import load_patterns as _load_patterns, PATTERNS_PATH


def _build_reverse_index(patterns):
    """Build keyword (lowercase) -> pattern_id index.
    For overlapping keywords, higher priority pattern wins; then longer keyword.
    NOTE: This index does NOT handle negative_keywords — use _match_vor_name_to_pattern
    with patterns list directly for full semantic matching.
    """
    index = {}
    # Sort patterns by priority DESC so higher priority patterns register first
    sorted_patterns = sorted(patterns, key=lambda p: p.get('priority', 10), reverse=True)
    for p in sorted_patterns:
        pid = p.get('id', '')
        for kw in p.get('keywords', []):
            kw_lower = kw.lower().strip()
            if kw_lower and kw_lower not in index:
                index[kw_lower] = pid
    return index


def _match_vor_name_to_pattern(vor_name_lower, patterns):
    """Find best pattern for a VOR name using priority, negative_keywords, and keyword length.

    Args:
        vor_name_lower: lowercased VOR position name
        patterns: list of pattern dicts from global_patterns.json

    Returns:
        pattern_id (str) of best match, or None if no match found
    """
    candidates = []

    for p in patterns:
        pat_id = p.get('id', '')
        keywords = p.get('keywords', [])
        negative = p.get('negative_keywords', [])
        priority = p.get('priority', 10)

        # Skip pattern if any negative keyword matches
        if any(neg.lower() in vor_name_lower for neg in negative if neg):
            continue

        # Find the best (longest) matching keyword
        best_kw_len = 0
        for kw in keywords:
            kw_lower = kw.lower().strip()
            if kw_lower and kw_lower in vor_name_lower:
                if len(kw_lower) > best_kw_len:
                    best_kw_len = len(kw_lower)

        if best_kw_len > 0:
            candidates.append((priority, best_kw_len, pat_id))

    if not candidates:
        return None

    # Sort: priority DESC, then keyword length DESC
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    return candidates[0][2]


def _build_id_to_group(patterns):
    """Build pattern_id -> group string mapping."""
    return {p['id']: p.get('group', '') for p in patterns if 'id' in p}


def _build_id_to_unit(patterns):
    """Build pattern_id -> canonical unit (m3/m2/m/count) mapping."""
    return {p['id']: p.get('unit', 'count') for p in patterns if 'id' in p}


# Cache for semantic matching results: (vor_name, frozenset(bim_labels)) -> matched_label
_SEMANTIC_CACHE = {}


def _semantic_match_vor_to_bim(vor_name: str, bim_categories: list) -> str:
    """Find the closest BIM category for a VOR position name via embeddings.

    Uses OpenAI text-embedding-3-small + cosine similarity.
    Caches results in _SEMANTIC_CACHE to avoid repeated API calls.
    Falls back to keyword matching if OpenAI is unavailable.

    Args:
        vor_name: VOR position name, e.g. "Кладка кирпичная наружная толщ. 510мм"
        bim_categories: list of BIM label strings to match against

    Returns:
        Best matching label from bim_categories, or None if no match found
    """
    if not bim_categories:
        return None

    cache_key = (vor_name, tuple(sorted(bim_categories)))
    if cache_key in _SEMANTIC_CACHE:
        return _SEMANTIC_CACHE[cache_key]

    result = None

    try:
        import math
        import openai

        def _cosine(a, b):
            dot = sum(x * y for x, y in zip(a, b))
            mag_a = math.sqrt(sum(x * x for x in a))
            mag_b = math.sqrt(sum(x * x for x in b))
            return dot / (mag_a * mag_b) if mag_a and mag_b else 0.0

        client = openai.OpenAI()
        texts = [vor_name] + list(bim_categories)
        response = client.embeddings.create(model="text-embedding-3-small", input=texts)
        embeddings = [r.embedding for r in response.data]

        vor_emb = embeddings[0]
        cat_embs = embeddings[1:]

        best_idx = max(range(len(cat_embs)), key=lambda i: _cosine(vor_emb, cat_embs[i]))
        best_sim = _cosine(vor_emb, cat_embs[best_idx])

        # Accept only if similarity exceeds minimum threshold
        if best_sim > 0.30:
            result = bim_categories[best_idx]

    except Exception:
        # Fallback: simple keyword overlap with BIM category names
        vor_lower = vor_name.lower()
        best_cat = None
        best_score = 0
        for cat in bim_categories:
            words = [w for w in cat.lower().split() if len(w) > 3]
            score = sum(1 for w in words if w in vor_lower)
            if score > best_score:
                best_score = score
                best_cat = cat
        if best_score > 0:
            result = best_cat

    _SEMANTIC_CACHE[cache_key] = result
    return result


# Cache for AI (gpt-4o-mini) matching results: vor_name -> matched_category
_AI_MATCH_CACHE = {}


def _ai_match_vor_to_bim(vor_name: str, bim_categories: list) -> str:
    """Find closest BIM category for a VOR position name using gpt-4o-mini.

    Uses OpenAI gpt-4o-mini for natural language semantic matching.
    Falls back to keyword overlap if OpenAI is unavailable.
    Caches results in _AI_MATCH_CACHE.

    Args:
        vor_name: VOR position name, e.g. "Кирпичная кладка стен из КСМ 190мм"
        bim_categories: list of BIM label strings to pick from

    Returns:
        Best matching label from bim_categories, or "NO_MATCH"
    """
    if not bim_categories:
        return "NO_MATCH"

    cache_key = (vor_name, tuple(sorted(bim_categories)))
    if cache_key in _AI_MATCH_CACHE:
        return _AI_MATCH_CACHE[cache_key]

    result = "NO_MATCH"

    try:
        import openai
        client = openai.OpenAI()

        cats_numbered = "\n".join(
            "{}.  {}".format(i + 1, cat) for i, cat in enumerate(bim_categories)
        )
        prompt = (
            "You are a construction BIM expert. "
            "Given a VOR (Bill of Quantities) position name, pick the BEST matching BIM category "
            "from the numbered list below. "
            "Reply ONLY with the exact category name as it appears in the list, nothing else. "
            "If none match, reply exactly: NO_MATCH\n\n"
            "VOR position: {vor}\n\n"
            "BIM categories:\n{cats}"
        ).format(vor=vor_name, cats=cats_numbered)

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=100,
            temperature=0.0,
        )
        answer = response.choices[0].message.content.strip()

        # Validate answer is one of the categories or NO_MATCH
        if answer in bim_categories:
            result = answer
        elif answer != "NO_MATCH":
            # Try fuzzy: strip number prefix if model echoed it
            answer_clean = answer.lstrip("0123456789. ").strip()
            if answer_clean in bim_categories:
                result = answer_clean
            else:
                result = "NO_MATCH"

    except Exception:
        # Fallback: simple keyword overlap with BIM category names
        vor_lower = vor_name.lower()
        best_cat = None
        best_score = 0
        for cat in bim_categories:
            words = [w for w in cat.lower().replace("_", " ").split() if len(w) > 3]
            score = sum(1 for w in words if w in vor_lower)
            if score > best_score:
                best_score = score
                best_cat = cat
        if best_score > 0:
            result = best_cat

    _AI_MATCH_CACHE[cache_key] = result
    return result


async def _fetch_bim_summary(revit_post, ctx):
    """Fetch bim_summary data (all semantic groups) from Revit model."""
    from ._constants import ALL_CATEGORIES, CAT_BATCHES
    from ._scan_engine import _build_batch_code
    from .bim_summary import _build_summary_from_catalog

    catalog_data = {}
    for batch in CAT_BATCHES:
        valid_batch = [c for c in batch if c in ALL_CATEGORIES]
        if not valid_batch:
            continue
        code = _build_batch_code(valid_batch, ALL_CATEGORIES, False)
        resp = await revit_post("/execute_code/", {"code": code}, ctx)
        if isinstance(resp, dict) and resp.get("status") == "success":
            raw = resp.get("output", "{}").strip()
            try:
                batch_result = json.loads(raw)
                catalog_data.update(batch_result)
            except Exception:
                pass

    if not catalog_data:
        return None

    patterns = _load_patterns()
    if not patterns:
        return None

    return _build_summary_from_catalog(catalog_data, patterns, "full")


def _extract_bim_vol_for_unit(bim_entry, vor_unit):
    """Pick correct volume value from bim_entry based on VOR unit string."""
    unit_lower = vor_unit.lower()
    if any(u in unit_lower for u in ('m3', 'м3', 'куб', 'кб.м', 'cubic')):
        return bim_entry.get('volume_m3', 0.0)
    if any(u in unit_lower for u in ('m2', 'м2', 'кв', 'sq', 'площ')):
        return bim_entry.get('area_m2', 0.0)
    if any(u in unit_lower for u in ('пог', 'п.м', 'lm', 'м пог')):
        return bim_entry.get('length_m', 0.0)
    # Fallback: use canonical unit from pattern
    pat_unit = bim_entry.get('pat_unit', 'count')
    if pat_unit == 'm3':
        return bim_entry.get('volume_m3', 0.0)
    if pat_unit == 'm2':
        return bim_entry.get('area_m2', 0.0)
    if pat_unit == 'm':
        return bim_entry.get('length_m', 0.0)
    return float(bim_entry.get('count', 0))


def _parse_vor_excel(file_path):
    """Parse VOR Excel file into list of dicts.

    Returns list of {"name": str, "unit": str, "volume": float}
    or dict with "error" key on failure.
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl required: pip install openpyxl"}

    if not os.path.exists(file_path):
        return {"error": "File not found: " + file_path}

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return {"error": "Cannot open Excel file: " + str(e)}

    # Pick sheet: prefer named sheets over first sheet
    ws = None
    for name in ("ВОР", "BIM", "Объёмы", "Объемы", "Volumes"):
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    # Keyword sets (lowercase)
    NAME_KW  = {"наименование", "название", "описание", "name", "работ", "позиция"}
    UNIT_KW  = {"ед", "единица", "unit", "ед.изм", "ед. изм", "ед.изм.", "изм"}
    VOL_KW   = {"объём", "объем", "количество", "кол-во", "volume", "qty", "кол.", "кол"}

    # Find header row (scan first 10 rows)
    header_row_idx = None
    col_name = col_unit = col_vol = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        # Check if any cell matches a name keyword
        row_strs = [(i, str(c).lower().strip() if c is not None else "") for i, c in enumerate(row)]
        has_name_kw = any(any(kw in cell_s for kw in NAME_KW) for _, cell_s in row_strs if cell_s)
        if not has_name_kw:
            continue

        # Found candidate header row — identify columns
        for col_idx, cell_s in row_strs:
            if cell_s == "":
                continue
            if any(kw in cell_s for kw in NAME_KW) and col_name is None:
                col_name = col_idx
            elif any(kw in cell_s for kw in UNIT_KW) and col_unit is None:
                col_unit = col_idx
            elif any(kw in cell_s for kw in VOL_KW) and col_vol is None:
                col_vol = col_idx

        if col_name is not None:
            header_row_idx = row_idx
            break

    if header_row_idx is None or col_name is None:
        return {"error": "Could not find header row in Excel file"}

    # Iterate data rows after header
    results = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # Extract name cell
        name_val = row[col_name] if col_name < len(row) else None
        if name_val is None or str(name_val).strip() == "":
            continue

        name_str = str(name_val).strip()

        # Skip separator/total rows
        name_lower = name_str.lower()
        if any(skip in name_lower for skip in ("итого", "итог", "total", "всего", "subtotal")):
            continue
        # Skip rows that are purely numeric
        try:
            float(name_str.replace(",", ".").replace(" ", ""))
            continue  # it's a number, skip
        except (ValueError, AttributeError):
            pass

        # Extract unit
        unit_str = ""
        if col_unit is not None and col_unit < len(row):
            uv = row[col_unit]
            unit_str = str(uv).strip() if uv is not None else ""

        # Extract volume
        vol_val = 0.0
        if col_vol is not None and col_vol < len(row):
            vv = row[col_vol]
            if vv is not None:
                try:
                    vol_val = float(str(vv).replace(",", ".").replace(" ", ""))
                except (ValueError, TypeError):
                    vol_val = 0.0

        results.append({"name": name_str, "unit": unit_str, "volume": vol_val})

    if not results:
        return {"error": "Excel file has no data rows"}

    return results


def register_vor_vs_bim_tools(mcp_server, revit_get, revit_post, revit_image):
    """Register VOR vs BIM comparison tools with semantic matching."""

    @mcp_server.tool()
    async def vor_vs_bim(
        vor_data: str = "[]",
        vor_file: str = "",
        tolerance_pct: float = 3.0,
        ctx: Context = None,
    ) -> dict:
        """Compare client VOR volumes with BIM model using semantic matching.

        vor_data: JSON string: [{"name": "Кирпичная кладка", "unit": "м3", "volume": 456.7}, ...]
        vor_file: path to .xlsx/.xls file with VOR table (takes priority over vor_data)
        tolerance_pct: acceptable deviation % (default 3.0, strict mode use 1.0)
        Returns: {matches, red_flags, missing_in_vor, summary}
        """
        # --- Resolve VOR items source ---
        vor_items = None

        # vor_file takes priority
        if vor_file and vor_file.strip():
            parsed = _parse_vor_excel(vor_file.strip())
            if isinstance(parsed, dict) and "error" in parsed:
                return parsed  # propagate parse error
            vor_items = parsed

        # Fall back to vor_data JSON
        if vor_items is None:
            raw_data = vor_data.strip() if vor_data else ""
            if not raw_data or raw_data == "[]":
                return {
                    "error": (
                        "No VOR data provided. "
                        "Pass vor_file='path/to/file.xlsx' or vor_data='[{...}]'"
                    )
                }
            err = validate_vor_data(raw_data)
            if err:
                return {"error": "Validation error: " + err}
            try:
                vor_items = json.loads(raw_data)
            except Exception as e:
                return {"error": "Invalid vor_data JSON: " + str(e)}
            if not isinstance(vor_items, list):
                return {"error": "vor_data must be a JSON array"}

        # Load semantic patterns
        patterns = _load_patterns()
        if not patterns:
            return {"error": "Could not load global_patterns.json", "path": PATTERNS_PATH}

        # Sort patterns by priority DESC once, for consistent matching
        patterns_sorted = sorted(patterns, key=lambda p: p.get('priority', 10), reverse=True)
        id_to_group = _build_id_to_group(patterns)
        id_to_unit = _build_id_to_unit(patterns)

        # Fetch BIM summary from Revit model
        summary = await _fetch_bim_summary(revit_post, ctx)
        if summary is None:
            return {"error": "Could not fetch BIM data from Revit"}

        # Build lookup: pattern_id -> bim_entry
        id_to_bim = {}
        for pid, group in id_to_group.items():
            parts = group.split(".", 1)
            top_key = parts[0]
            sub_key = parts[1] if len(parts) > 1 else "other"
            if top_key in summary and sub_key in summary[top_key]:
                grp = summary[top_key][sub_key]
                id_to_bim[pid] = {
                    "volume_m3": grp.get("total_volume_m3", 0.0),
                    "area_m2":   grp.get("total_area_m2",   0.0),
                    "length_m":  grp.get("total_length_m",  0.0),
                    "count":     grp.get("total_count",      0),
                    "label":     grp.get("label",           ""),
                    "pat_unit":  id_to_unit.get(pid, "count"),
                    "group":     group,
                    "bim_types": [b.get("type", "") for b in grp.get("breakdown", [])[:5]],
                }

        # Build reverse label->pid lookup for semantic fallback
        labels_to_pid = {
            bim_entry["label"]: pid
            for pid, bim_entry in id_to_bim.items()
            if bim_entry.get("label")
        }
        bim_label_list = list(labels_to_pid.keys())

        # Match VOR items to BIM semantic groups
        matches = []
        red_flags = []
        matched_patterns = set()

        for item in vor_items:
            name = item.get("name", "")
            vor_vol = float(item.get("volume", 0) or 0)
            unit = item.get("unit", "")
            name_lower = name.lower().strip()

            # Step 1: keyword matching
            pattern_id = _match_vor_name_to_pattern(name_lower, patterns_sorted)
            match_method = "keyword" if pattern_id else "none"

            # Step 2: AI (gpt-4o-mini) fallback when keyword matching fails
            if pattern_id is None and bim_label_list:
                ai_label = _ai_match_vor_to_bim(name, bim_label_list)
                if ai_label and ai_label != "NO_MATCH" and ai_label in labels_to_pid:
                    pattern_id = labels_to_pid[ai_label]
                    match_method = "ai"

            # Step 3: embedding semantic fallback when both keyword and AI fail
            if pattern_id is None and bim_label_list:
                best_label = _semantic_match_vor_to_bim(name, bim_label_list)
                if best_label and best_label in labels_to_pid:
                    pattern_id = labels_to_pid[best_label]
                    match_method = "semantic"

            bim_entry = id_to_bim.get(pattern_id) if pattern_id else None

            bim_vol = None
            if bim_entry is not None:
                bim_vol = _extract_bim_vol_for_unit(bim_entry, unit)
                matched_patterns.add(pattern_id)

            entry = {
                "name":            name,
                "unit":            unit,
                "vor_volume":      vor_vol,
                "bim_volume":      bim_vol,
                "matched_pattern": pattern_id,
                "bim_label":       bim_entry["label"] if bim_entry else None,
                "match_method":    match_method,
            }

            if bim_vol is None:
                entry["status"] = "no_bim_match"
                matches.append(entry)
            elif vor_vol == 0:
                entry["status"] = "zero_in_vor"
                entry["diff_pct"] = None
                red_flags.append(entry)
            else:
                diff_pct = abs(vor_vol - bim_vol) / vor_vol * 100
                entry["diff_pct"] = round(diff_pct, 2)
                if diff_pct > tolerance_pct:
                    entry["status"] = "red_flag"
                    red_flags.append(entry)
                else:
                    entry["status"] = "ok"
                    matches.append(entry)

        # missing_in_vor: BIM patterns with volume but not mentioned in VOR
        missing_in_vor = []
        for pid, bim_entry in id_to_bim.items():
            if pid in matched_patterns:
                continue
            bim_vol_check = (
                bim_entry.get("volume_m3", 0) or
                bim_entry.get("area_m2", 0) or
                bim_entry.get("length_m", 0) or
                bim_entry.get("count", 0)
            )
            if bim_vol_check and bim_vol_check > 0:
                missing_in_vor.append({
                    "pattern_id":    pid,
                    "label":         bim_entry["label"],
                    "group":         bim_entry["group"],
                    "bim_volume_m3": bim_entry.get("volume_m3", 0),
                    "bim_area_m2":   bim_entry.get("area_m2",   0),
                    "bim_length_m":  bim_entry.get("length_m",  0),
                    "count":         bim_entry.get("count",     0),
                })

        missing_in_vor.sort(
            key=lambda x: -(x.get("bim_volume_m3") or x.get("bim_area_m2") or 0)
        )

        return {
            "matches":        matches,
            "red_flags":      red_flags,
            "missing_in_vor": missing_in_vor,
            "summary": {
                "total_vor":      len(vor_items),
                "ok":             len([m for m in matches if m.get("status") == "ok"]),
                "red_flags":      len(red_flags),
                "no_match":       len([m for m in matches if m.get("status") == "no_bim_match"]),
                "missing":        len(missing_in_vor),
                "tolerance_pct":  tolerance_pct,
                "tolerance_used": tolerance_pct,
                "patterns_loaded": len(patterns),
                "source":         "excel:" + vor_file if vor_file else "json",
            },
        }

    @mcp_server.tool()
    async def vor_vs_bim_file(
        vor_file: str,
        tolerance_pct: float = 3.0,
        ctx: Context = None,
    ) -> dict:
        """Load VOR from Excel file and compare with BIM model.

        Convenience wrapper around vor_vs_bim — just pass the path to the .xlsx file.
        vor_file: absolute or relative path to Excel file with VOR table
        tolerance_pct: acceptable deviation % (default 3.0, strict mode use 1.0)
        Returns: {matches, red_flags, missing_in_vor, summary}

        Excel format: the tool auto-detects columns for name/unit/volume.
        Supported sheet names: ВОР, BIM, Объёмы (or first sheet).
        """
        if not vor_file or not vor_file.strip():
            return {"error": "vor_file path is required"}

        # Delegate to vor_vs_bim with vor_file set
        return await vor_vs_bim(
            vor_data="[]",
            vor_file=vor_file,
            tolerance_pct=tolerance_pct,
            ctx=ctx,
        )

    @mcp_server.tool()
    async def vor_excel_preview(vor_file: str) -> dict:
        """Preview Excel VOR file parsing — shows detected columns and first 10 rows.

        Useful to verify auto-detection before running full vor_vs_bim comparison.
        vor_file: path to .xlsx file
        Returns: {columns_detected, rows_preview, total_rows}
        """
        if not vor_file or not vor_file.strip():
            return {"error": "vor_file path is required"}

        result = _parse_vor_excel(vor_file.strip())
        if isinstance(result, dict) and "error" in result:
            return result

        return {
            "total_rows": len(result),
            "rows_preview": result[:10],
            "columns_detected": {
                "name": "auto",
                "unit": "auto",
                "volume": "auto",
            },
            "message": (
                "Detected {} positions. Use vor_vs_bim_file to run full comparison.".format(
                    len(result)
                )
            ),
        }
