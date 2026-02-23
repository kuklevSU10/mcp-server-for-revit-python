# -*- coding: utf-8 -*-
"""BIM data export to Excel."""
import json
import os
from datetime import datetime
from mcp.server.fastmcp import Context

# ── Revit category names for bim_catalog detection ────────────────────────────
_REVIT_CATS = {
    'Walls', 'Floors', 'Roofs', 'Ceilings', 'Columns', 'StructuralFraming',
    'StructuralFoundation', 'Doors', 'Windows', 'Furniture', 'GenericModel',
    'Ducts', 'Pipes', 'MechanicalEquipment', 'ElectricalEquipment',
    'LightingFixtures', 'CableTray', 'Conduit', 'Ramps', 'Stairs',
}

# ── Style helpers ─────────────────────────────────────────────────────────────

def _header_font():
    from openpyxl.styles import Font
    return Font(bold=True, size=11)


def _group_fill(color_hex='DDEBF6'):
    from openpyxl.styles import PatternFill
    return PatternFill(fill_type='solid', fgColor=color_hex)


def _thin_border():
    from openpyxl.styles import Border, Side
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)


def _set_col_widths(ws, widths):
    """widths: list of (col_letter, width)"""
    for col, w in widths:
        ws.column_dimensions[col].width = w


def _write_header_row(ws, row, headers, fill=None):
    from openpyxl.styles import Font, Alignment
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if fill:
            cell.fill = fill


# ── Format detector ───────────────────────────────────────────────────────────

def _detect_data_type(data):
    """Detect which BIM tool produced this data."""
    if not isinstance(data, dict):
        return 'generic'
    keys = set(data.keys())
    # bim_summary: has structural / architectural / mep groups
    if keys & {'structural', 'architectural', 'mep'}:
        return 'bim_summary'
    # vor_vs_bim: has red_flags + matches
    if 'red_flags' in keys and 'matches' in keys:
        return 'vor_vs_bim'
    # bim_vor_generate: has positions + total_positions
    if 'positions' in keys and 'total_positions' in keys:
        return 'bim_vor_generate'
    # bim_catalog: keys match Revit category names
    if keys & _REVIT_CATS:
        return 'bim_catalog'
    return 'generic'


# ── Sheet writers ─────────────────────────────────────────────────────────────

def _write_summary_sheet(ws, data):
    """Write bim_summary result to worksheet.

    bim_summary format:
      {group_name: {sub_key: {label, total_count, total_volume_m3, total_area_m2,
                               total_length_m, breakdown: [...]}}}
      group_name: 'structural', 'architectural', 'mep'
      sub_key:    'monolith', 'masonry_aac', etc. (pattern group suffix)
    """
    from openpyxl.styles import Font, Alignment

    today = datetime.now().strftime('%Y-%m-%d')
    ws.cell(row=1, column=1, value='BIM Сводка — {}'.format(today)).font = Font(bold=True, size=13)

    headers = ['Группа', 'Паттерн', 'Наименование', 'Объём м3', 'Площадь м2', 'Длина м', 'Кол-во']
    _write_header_row(ws, 3, headers, fill=_group_fill('BDD7EE'))
    _set_col_widths(ws, [('A', 18), ('B', 20), ('C', 35), ('D', 12), ('E', 12), ('F', 12), ('G', 10)])

    row = 4
    group_labels = {
        'structural': 'Конструктив (Structural)',
        'architectural': 'Архитектура (Architectural)',
        'mep': 'Инженерия (MEP)',
    }

    total_vol = total_area = total_length = total_count = 0.0

    for group_key in ('structural', 'architectural', 'mep'):
        group_data = data.get(group_key)
        if not group_data or not isinstance(group_data, dict):
            continue

        # Group header row
        group_label = group_labels.get(group_key, group_key.upper())
        group_cell = ws.cell(row=row, column=1, value=group_label)
        group_cell.font = Font(bold=True)
        group_fill = _group_fill('DDEBF6')
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill = group_fill
        row += 1

        # Iterate sub-patterns within the group
        # group_data: {sub_key: {label, total_count, total_volume_m3, ...}}
        for sub_key, pdata in group_data.items():
            if not isinstance(pdata, dict):
                continue
            label = pdata.get('label', sub_key)
            vol = float(pdata.get('total_volume_m3', 0) or 0)
            area = float(pdata.get('total_area_m2', 0) or 0)
            length = float(pdata.get('total_length_m', 0) or 0)
            count = int(pdata.get('total_count', 0) or 0)

            ws.cell(row=row, column=1, value=group_key)
            ws.cell(row=row, column=2, value=sub_key)
            ws.cell(row=row, column=3, value=label)
            ws.cell(row=row, column=4, value=round(vol, 2))
            ws.cell(row=row, column=5, value=round(area, 2))
            ws.cell(row=row, column=6, value=round(length, 2))
            ws.cell(row=row, column=7, value=count)

            total_vol += vol
            total_area += area
            total_length += length
            total_count += count
            row += 1

    # Totals row
    row += 1
    ws.cell(row=row, column=1, value='ИТОГО').font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(total_vol, 2)).font = Font(bold=True)
    ws.cell(row=row, column=5, value=round(total_area, 2)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(total_length, 2)).font = Font(bold=True)
    ws.cell(row=row, column=7, value=int(total_count)).font = Font(bold=True)


def _write_vor_comparison_sheet(ws, data):
    """Write vor_vs_bim result to worksheet."""
    from openpyxl.styles import PatternFill

    headers = ['Позиция ВОР', 'Единица', 'ВОР объём', 'BIM объём', 'Расхождение %', 'Статус']
    _write_header_row(ws, 1, headers, fill=_group_fill('BDD7EE'))
    _set_col_widths(ws, [('A', 35), ('B', 10), ('C', 14), ('D', 14), ('E', 16), ('F', 14)])

    fill_red = PatternFill(fill_type='solid', fgColor='FFCCCC')
    fill_green = PatternFill(fill_type='solid', fgColor='CCFFCC')
    fill_yellow = PatternFill(fill_type='solid', fgColor='FFFF99')

    row = 2
    ok_count = red_count = no_match_count = 0

    def write_row(item, fill):
        ws.cell(row=row, column=1, value=item.get('name', ''))
        ws.cell(row=row, column=2, value=item.get('unit', ''))
        ws.cell(row=row, column=3, value=item.get('vor_volume', ''))
        ws.cell(row=row, column=4, value=item.get('bim_volume', ''))
        ws.cell(row=row, column=5, value=item.get('diff_pct', ''))
        ws.cell(row=row, column=6, value=item.get('status', ''))
        if fill:
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = fill

    for item in data.get('matches', []):
        st = item.get('status', '')
        fill = fill_green if st == 'ok' else fill_yellow
        write_row(item, fill)
        if st == 'ok':
            ok_count += 1
        else:
            no_match_count += 1
        row += 1

    for item in data.get('red_flags', []):
        write_row(item, fill_red)
        red_count += 1
        row += 1

    # Summary
    row += 1
    from openpyxl.styles import Font
    ws.cell(row=row, column=1, value='OK: {}'.format(ok_count)).font = Font(bold=True)
    ws.cell(row=row, column=2, value='Red flags: {}'.format(red_count)).font = Font(bold=True, color='FF0000')
    ws.cell(row=row, column=3, value='Нет совпадения: {}'.format(no_match_count)).font = Font(bold=True)


def _write_vor_positions_sheet(ws, data):
    """Write bim_vor_generate result — ready VOR for tender."""
    headers = ['№', 'Наименование работ', 'Единица', 'Объём', 'Группа', 'Тип BIM']
    _write_header_row(ws, 1, headers, fill=_group_fill('BDD7EE'))
    _set_col_widths(ws, [('A', 5), ('B', 45), ('C', 10), ('D', 12), ('E', 18), ('F', 25)])

    from openpyxl.styles import Alignment
    positions = data.get('positions', [])
    for i, pos in enumerate(positions, 1):
        r = i + 1
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=pos.get('name', '')).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=3, value=pos.get('unit', ''))
        ws.cell(row=r, column=4, value=pos.get('volume', pos.get('value', '')))
        ws.cell(row=r, column=5, value=pos.get('group', ''))
        # bim_vor_generate uses 'bim_types' (list), take first; fallback 'bim_type' or 'type'
        bim_types = pos.get('bim_types', [])
        if isinstance(bim_types, list):
            bim_type_str = bim_types[0] if bim_types else pos.get('bim_type', pos.get('type', ''))
        else:
            bim_type_str = str(bim_types)
        ws.cell(row=r, column=6, value=bim_type_str)

    # Footer
    from openpyxl.styles import Font
    footer_row = len(positions) + 3
    ws.cell(row=footer_row, column=1,
            value='Итого позиций: {}'.format(data.get('total_positions', len(positions)))).font = Font(bold=True)


def _write_catalog_sheet(ws, data):
    """Write bim_catalog result."""
    headers = ['Категория', 'Кол-во типов', 'Общий объём м3', 'Общая площадь м2']
    _write_header_row(ws, 1, headers, fill=_group_fill('BDD7EE'))
    _set_col_widths(ws, [('A', 25), ('B', 14), ('C', 16), ('D', 16)])

    row = 2
    for cat_name, cat_data in data.items():
        if not isinstance(cat_data, dict):
            continue
        ws.cell(row=row, column=1, value=cat_name)
        # count of types or total count
        types_val = cat_data.get('types')
        if isinstance(types_val, list):
            types_val = len(types_val)
        elif types_val is None:
            types_val = cat_data.get('type_count', cat_data.get('count', ''))
        ws.cell(row=row, column=2, value=types_val)
        ws.cell(row=row, column=3, value=cat_data.get('volume_m3', cat_data.get('total_volume_m3', '')))
        ws.cell(row=row, column=4, value=cat_data.get('area_m2', cat_data.get('total_area_m2', '')))
        row += 1


def _write_generic_sheet(ws, data):
    """Write any JSON dict as key-value table."""
    from openpyxl.styles import Font
    _write_header_row(ws, 1, ['Ключ', 'Значение'], fill=_group_fill('BDD7EE'))
    _set_col_widths(ws, [('A', 30), ('B', 50)])

    row = 2
    for key, value in data.items():
        ws.cell(row=row, column=1, value=str(key))
        if isinstance(value, (dict, list)):
            ws.cell(row=row, column=2, value=json.dumps(value, ensure_ascii=False)[:500])
        else:
            ws.cell(row=row, column=2, value=str(value) if value is not None else '')
        row += 1


# ── Registration ──────────────────────────────────────────────────────────────

def register_bim_export_tools(mcp_server, revit_get, revit_post, revit_image):
    """Register BIM Excel export tools."""

    @mcp_server.tool()
    async def bim_export_excel(
        data: str,
        output_path: str = '',
        title: str = 'BIM Report',
        ctx: Context = None,
    ) -> dict:
        """Export BIM tool results to Excel (.xlsx).

        data: JSON string with BIM tool result
              (from bim_summary, bim_catalog, vor_vs_bim, bim_vor_generate)
        output_path: path to save .xlsx file
                     (default: C:/Users/kuklev.d.s/clawd/exports/bim_TIMESTAMP.xlsx)
        title: sheet title (max 31 chars, Excel limit)

        Auto-detects data format:
        - bim_summary result      -> 'Сводка' sheet with groups/subgroups
        - bim_catalog result      -> categories sheet
        - vor_vs_bim result       -> 'ВОР vs BIM' sheet, red flags in red
        - bim_vor_generate result -> 'Позиции ВОР' sheet ready for tender
        """
        try:
            parsed = json.loads(data)
        except Exception as e:
            return {"error": "Invalid JSON data: " + str(e)}

        # Determine output path
        if not output_path:
            import time
            ts = int(time.time())
            exports_dir = r'C:\Users\kuklev.d.s\clawd\exports'
            os.makedirs(exports_dir, exist_ok=True)
            output_path = os.path.join(exports_dir, 'bim_{}.xlsx'.format(ts))

        try:
            import openpyxl
        except ImportError:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit

        data_type = _detect_data_type(parsed)

        if data_type == 'bim_summary':
            _write_summary_sheet(ws, parsed)
        elif data_type == 'vor_vs_bim':
            _write_vor_comparison_sheet(ws, parsed)
        elif data_type == 'bim_vor_generate':
            _write_vor_positions_sheet(ws, parsed)
        elif data_type == 'bim_catalog':
            _write_catalog_sheet(ws, parsed)
        else:
            _write_generic_sheet(ws, parsed)

        try:
            ws.freeze_panes = 'A2'
        except Exception:
            pass

        wb.save(output_path)
        return {
            "status": "ok",
            "path": output_path,
            "data_type": data_type,
            "sheets": [s.title for s in wb.worksheets],
        }
